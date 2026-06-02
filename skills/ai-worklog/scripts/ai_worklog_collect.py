from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import re
import subprocess
from pathlib import Path
from typing import Any

SENSITIVE_RE = re.compile(r"(?i)(password|passwd|pwd|token|secret|cookie|authorization|access[_-]?token|refresh[_-]?token|private[_-]?key|api[_-]?key)\s*[:=]\s*[^\s,;]+")
DSN_RE = re.compile(r"(?i)(postgres(?:ql)?://[^\s]+)")


def redact(text: Any) -> str:
    value = "" if text is None else str(text)
    value = SENSITIVE_RE.sub(lambda m: m.group(1) + "=[REDACTED]", value)
    value = DSN_RE.sub("[REDACTED_DSN]", value)
    return value


def shorten(text: Any, limit: int = 180) -> str:
    value = redact(text).replace("\r", " ").replace("\n", " ").strip()
    value = re.sub(r"\s+", " ", value)
    return value if len(value) <= limit else value[: limit - 3] + "..."


def parse_date(value: str) -> dt.date:
    today = dt.datetime.now().astimezone().date()
    if value == "today":
        return today
    if value == "yesterday":
        return today - dt.timedelta(days=1)
    return dt.date.fromisoformat(value)


def parse_ts(value: Any) -> dt.datetime | None:
    if not value:
        return None
    if isinstance(value, (int, float)):
        return dt.datetime.fromtimestamp(value / 1000, tz=dt.timezone.utc).astimezone()
    text = str(value)
    try:
        if text.endswith("Z"):
            text = text[:-1] + "+00:00"
        return dt.datetime.fromisoformat(text).astimezone()
    except ValueError:
        return None


def in_day(ts: dt.datetime | None, day: dt.date) -> bool:
    return bool(ts and ts.astimezone().date() == day)


def read_jsonl(path: Path):
    try:
        with path.open("r", encoding="utf-8") as handle:
            for line in handle:
                line = line.strip()
                if not line:
                    continue
                try:
                    yield json.loads(line)
                except json.JSONDecodeError:
                    continue
    except OSError:
        return


def codex_home() -> Path:
    return Path(os.environ.get("CODEX_HOME") or Path.home() / ".codex").expanduser()


def claude_home() -> Path:
    return Path(os.environ.get("CLAUDE_CONFIG_DIR") or Path.home() / ".claude").expanduser()


def collect_codex(day: dt.date) -> tuple[list[dict[str, Any]], list[Path], set[Path]]:
    home = codex_home()
    session_dir = home / "sessions" / f"{day:%Y}" / f"{day:%m}" / f"{day:%d}"
    rows: list[dict[str, Any]] = []
    files: list[Path] = []
    cwds: set[Path] = set()
    if not session_dir.exists():
        return rows, files, cwds
    for path in sorted(session_dir.glob("*.jsonl")):
        files.append(path)
        current: dict[str, Any] | None = None
        last_agent = ""
        last_ts: dt.datetime | None = None
        cwd = None
        for obj in read_jsonl(path):
            ts = parse_ts(obj.get("timestamp"))
            if obj.get("type") == "session_meta":
                cwd_text = ((obj.get("payload") or {}).get("cwd"))
                if cwd_text:
                    cwd = Path(cwd_text)
                    cwds.add(cwd)
            payload = obj.get("payload") or {}
            typ = obj.get("type")
            ptyp = payload.get("type")
            if typ == "event_msg" and ptyp == "user_message" and in_day(ts, day):
                if current:
                    rows.append(finish_row(current, last_ts, last_agent, path))
                current = {
                    "tool": "Codex",
                    "start": ts,
                    "end": ts,
                    "request": shorten(payload.get("message"), 220),
                    "result": "",
                    "cwd": str(cwd) if cwd else "",
                    "evidence": str(path),
                    "used_tool": False,
                }
                last_agent = ""
                last_ts = ts
            elif typ == "event_msg" and ptyp == "agent_message" and current:
                last_agent = shorten(payload.get("message"), 260)
                last_ts = ts or last_ts
            elif typ == "event_msg" and ptyp == "task_complete" and current:
                rows.append(finish_row(current, ts or last_ts, last_agent, path))
                current = None
                last_agent = ""
                last_ts = None
            elif typ == "response_item" and current:
                if payload.get("type") == "function_call":
                    current["used_tool"] = True
                last_ts = ts or last_ts
        if current:
            rows.append(finish_row(current, last_ts, last_agent, path))
    return rows, files, cwds


def finish_row(current: dict[str, Any], end: dt.datetime | None, result: str, path: Path) -> dict[str, Any]:
    current["end"] = end or current.get("start")
    current["result"] = result or current.get("result") or "未提取到最终回复"
    current["evidence"] = str(path)
    return current


def is_real_claude_user_content(content: Any) -> str | None:
    if isinstance(content, str):
        if content.startswith("<command-message>"):
            return content
        return content
    if isinstance(content, list):
        texts = []
        for item in content:
            if not isinstance(item, dict):
                continue
            if item.get("type") == "text":
                text = item.get("text") or ""
                if text.startswith("Base directory for this skill:"):
                    continue
                if "# PostgreSQL Query" in text and "ARGUMENTS:" in text:
                    continue
                texts.append(text)
            elif item.get("type") == "tool_result":
                continue
        return " ".join(texts).strip() or None
    return None


def collect_claude(day: dt.date) -> tuple[list[dict[str, Any]], list[Path], set[Path]]:
    home = claude_home()
    rows: list[dict[str, Any]] = []
    files: list[Path] = []
    cwds: set[Path] = set()
    projects = home / "projects"
    if not projects.exists():
        return rows, files, cwds
    for path in sorted(projects.rglob("*.jsonl")):
        try:
            if dt.datetime.fromtimestamp(path.stat().st_mtime).date() != day:
                continue
        except OSError:
            continue
        files.append(path)
        current: dict[str, Any] | None = None
        last_text = ""
        last_ts: dt.datetime | None = None
        for obj in read_jsonl(path):
            ts = parse_ts(obj.get("timestamp"))
            if not in_day(ts, day):
                continue
            if obj.get("cwd"):
                cwd = Path(str(obj.get("cwd")))
                cwds.add(cwd)
            if obj.get("type") == "user" and (obj.get("message") or {}).get("role") == "user":
                text = is_real_claude_user_content((obj.get("message") or {}).get("content"))
                if not text:
                    continue
                if current:
                    rows.append(finish_row(current, last_ts, last_text, path))
                current = {
                    "tool": "Claude",
                    "start": ts,
                    "end": ts,
                    "request": shorten(text, 220),
                    "result": "",
                    "cwd": str(obj.get("cwd") or ""),
                    "evidence": str(path),
                    "used_tool": False,
                }
                last_text = ""
                last_ts = ts
            elif obj.get("type") == "assistant" and current:
                parts = []
                for item in ((obj.get("message") or {}).get("content") or []):
                    if isinstance(item, dict) and item.get("type") == "text":
                        parts.append(item.get("text") or "")
                    elif isinstance(item, dict) and item.get("type") == "tool_use":
                        parts.append("工具调用: " + str(item.get("name") or "unknown"))
                if parts:
                    last_text = shorten(" | ".join(parts), 260)
                last_ts = ts or last_ts
            elif obj.get("type") == "system" and obj.get("subtype") == "turn_duration" and current:
                last_ts = ts or last_ts
                rows.append(finish_row(current, last_ts, last_text, path))
                current = None
                last_text = ""
        if current:
            rows.append(finish_row(current, last_ts, last_text, path))
    return rows, files, cwds


def is_git_repo(path: Path) -> bool:
    return (path / ".git").exists()


def discover_repos(explicit: list[str], roots: list[str], cwds: set[Path]) -> list[Path]:
    candidates: set[Path] = set()
    for raw in explicit:
        p = Path(raw).expanduser().resolve()
        if is_git_repo(p):
            candidates.add(p)
    for cwd in cwds:
        try:
            p = cwd.expanduser().resolve()
        except OSError:
            continue
        for parent in [p, *p.parents]:
            if is_git_repo(parent):
                candidates.add(parent)
                break
    for raw in roots:
        root = Path(raw).expanduser().resolve()
        if is_git_repo(root):
            candidates.add(root)
        if root.exists():
            for child in root.glob("*"):
                if child.is_dir() and is_git_repo(child):
                    candidates.add(child.resolve())
                elif child.is_dir():
                    for grand in child.glob("*"):
                        if grand.is_dir() and is_git_repo(grand):
                            candidates.add(grand.resolve())
    return sorted(candidates, key=lambda p: str(p).lower())


def run_git(repo: Path, args: list[str]) -> str:
    try:
        result = subprocess.run(["git", *args], cwd=repo, text=True, capture_output=True, timeout=20)
        return redact((result.stdout or "") + (("\n" + result.stderr) if result.stderr else "")).strip()
    except Exception as exc:
        return f"git_error: {exc}"


def collect_git(day: dt.date, repos: list[Path]) -> list[dict[str, Any]]:
    start = f"{day.isoformat()} 00:00:00"
    end = f"{(day + dt.timedelta(days=1)).isoformat()} 00:00:00"
    rows = []
    for repo in repos:
        log = run_git(repo, ["log", f"--since={start}", f"--until={end}", "--date=iso", "--pretty=format:%h\t%ad\t%an\t%s"])
        status = run_git(repo, ["status", "--short"])
        rows.append({"repo": str(repo), "commits": log, "status": status})
    return rows


def duration_minutes(row: dict[str, Any]) -> float:
    start = row.get("start")
    end = row.get("end")
    if not isinstance(start, dt.datetime) or not isinstance(end, dt.datetime):
        return 0.0
    minutes = max(0.0, (end - start).total_seconds() / 60.0)
    return round(minutes, 1)


def estimate_multiplier(row: dict[str, Any]) -> float:
    text = f"{row.get('request', '')} {row.get('result', '')}".lower()
    if any(k in text for k in ["commit", "git", "??", "??", "??", "mapper", "??", "bug"]):
        return 3.5
    if any(k in text for k in ["sql", "postgres", "pg", "???", "??", "??sql"]):
        return 3.0
    if any(k in text for k in ["skill", "??", "??", "???"]):
        return 3.0
    if any(k in text for k in ["??", "??", "sheet", "??", "??"]):
        return 2.5
    if len(text) < 120 and not row.get("used_tool"):
        return 1.5
    return 2.5


def effective_ai_minutes(row: dict[str, Any]) -> float:
    minutes = duration_minutes(row)
    if minutes > 30 and not row.get("used_tool"):
        return 30.0
    return minutes


def estimate_work_minutes(row: dict[str, Any]) -> float:
    return round(effective_ai_minutes(row) * estimate_multiplier(row), 1)


def row_to_jsonable(row: dict[str, Any]) -> dict[str, Any]:
    out = dict(row)
    for key in ("start", "end"):
        if isinstance(out.get(key), dt.datetime):
            out[key] = out[key].strftime("%Y-%m-%d %H:%M:%S")
    out["ai_active_minutes"] = duration_minutes(row)
    out["estimated_work_minutes"] = estimate_work_minutes(row)
    out["estimate_multiplier"] = estimate_multiplier(row)
    out["used_tool"] = bool(row.get("used_tool"))
    return out


def markdown_table(headers: list[str], rows: list[list[Any]]) -> str:
    lines = ["| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for row in rows:
        lines.append("| " + " | ".join(str(x).replace("|", "\\|") for x in row) + " |")
    return "\n".join(lines)


def render_markdown(data: dict[str, Any]) -> str:
    conv = data["conversations"]
    sources = data["sources"]
    git_rows = data["git"]
    total_ai = round(sum(row["ai_active_minutes"] for row in conv), 1)
    total_work = round(sum(row["estimated_work_minutes"] for row in conv), 1)
    source_rows = [[s["source"], s["path"], s["count"], s["status"]] for s in sources]
    conv_rows = []
    for idx, row in enumerate(conv, 1):
        conv_rows.append([
            idx,
            row["tool"],
            f"{row['start'][-8:-3]}-{row['end'][-8:-3]}",
            row["request"],
            row["result"],
            row["ai_active_minutes"],
            Path(row["evidence"]).name,
        ])
    git_table_rows = []
    for row in git_rows:
        commits = row["commits"] if row["commits"] else "无当天 commit"
        status = "有未提交变更" if row["status"] else "干净"
        git_table_rows.append([row["repo"], shorten(commits, 180), status])
    parts = [f"# AI Worklog {data['date']}", "", "## 数据来源", markdown_table(["来源", "路径", "记录数", "状态"], source_rows), "", "## AI 对话明细", markdown_table(["序号", "AI工具", "时间", "用户请求摘要", "AI处理结果", "AI活跃耗时(分钟)", "证据"], conv_rows), "", "## Git 证据", markdown_table(["仓库", "当天提交", "状态"], git_table_rows), "", "## 耗时汇总", markdown_table(["指标", "分钟", "小时"], [["AI活跃耗时", total_ai, round(total_ai / 60, 2)], ["估算真实工作耗时", total_work, round(total_work / 60, 2)]]), ""]
    return "\n".join(parts)




def default_excel_path() -> Path:
    return Path(__file__).resolve().parents[1] / "data" / "ai_worklog.xlsx"


def resolve_excel_path(value: str | None) -> Path:
    if not value:
        return default_excel_path()
    return Path(value).expanduser().resolve()


def cn(value: str) -> str:
    return value.encode("ascii").decode("unicode_escape")



def default_preview_path() -> Path:
    return Path(__file__).resolve().parents[1] / "data" / "ai_worklog_preview.xlsx"


def resolve_preview_path(value: str | None) -> Path:
    if not value:
        return default_preview_path()
    return Path(value).expanduser().resolve()


def default_config_path() -> Path:
    return Path(__file__).resolve().parents[1] / "data" / "privacy_config.local.jsonc"


def default_state_path() -> Path:
    return Path(__file__).resolve().parents[1] / "data" / "ai_worklog_state.local.json"


def resolve_state_path(value: str | None) -> Path:
    if not value:
        return default_state_path()
    return Path(value).expanduser().resolve()


def default_privacy_config_text() -> str:
    return cn(r'''{
  // \u9690\u79c1\u9ed1\u540d\u5355\uff1a\u547d\u4e2d\u4ee5\u4e0b\u8bcd\u7684 AI \u5bf9\u8bdd\u4f1a\u88ab\u6807\u8bb0\u4e3a\u9ad8\u98ce\u9669\uff0c\u9ed8\u8ba4\u4e0d\u7eb3\u5165\u65e5\u62a5\u3002\u53ef\u6309\u56e2\u961f\u9700\u6c42\u81ea\u884c\u589e\u5220\u3002
  "privacy_blacklist": ["\u5de5\u8d44", "\u85aa\u916c", "\u5956\u91d1", "\u7ee9\u6548", "\u5410\u69fd", "\u79bb\u804c", "\u5bb6\u5ead", "\u79c1\u4eba", "\u8eab\u4efd\u8bc1", "\u94f6\u884c\u5361", "\u5408\u540c\u91d1\u989d", "token", "cookie", "password", "secret"],

  // \u9879\u76ee\u767d\u540d\u5355\uff1a\u975e\u7a7a\u65f6\uff0c\u53ea\u6709\u6458\u8981\u3001\u8def\u5f84\u6216\u8bc1\u636e\u4e2d\u547d\u4e2d\u8fd9\u4e9b\u5173\u952e\u8bcd\u7684\u8bb0\u5f55\u624d\u4f1a\u9ed8\u8ba4\u7eb3\u5165\u3002\u7a7a\u6570\u7ec4\u8868\u793a\u4e0d\u542f\u7528\u767d\u540d\u5355\u9650\u5236\u3002
  "project_whitelist": [],

  // \u9879\u76ee\u9ed1\u540d\u5355\uff1a\u6458\u8981\u3001\u8def\u5f84\u6216\u8bc1\u636e\u4e2d\u547d\u4e2d\u8fd9\u4e9b\u5173\u952e\u8bcd\u65f6\uff0c\u9ed8\u8ba4\u6392\u9664\u3002\u9002\u5408\u586b\u5199 personal/private/family \u7b49\u975e\u5de5\u4f5c\u76ee\u5f55\u6216\u9879\u76ee\u540d\u3002
  "project_blacklist": ["personal", "private", "family"],

  // \u6392\u9664 AI \u5de5\u5177\uff1a\u53ef\u586b\u5199 Codex\u3001Claude \u7b49\u3002\u547d\u4e2d\u540e\u8be5\u5de5\u5177\u7684\u8bb0\u5f55\u9ed8\u8ba4\u4e0d\u7eb3\u5165\u3002
  "exclude_tools": [],

  // \u6392\u9664\u65f6\u95f4\u6bb5\uff1a\u7528\u4e8e\u6392\u9664\u5348\u4f11\u3001\u4e0b\u73ed\u540e\u3001\u79c1\u4eba\u4f7f\u7528\u7b49\u65f6\u95f4\u3002\u683c\u5f0f\u4e3a HH:mm\uff0c\u53ef\u914d\u591a\u6bb5\u3002
  "exclude_time_ranges": [],

  // \u9ed8\u8ba4\u7eb3\u5165\u7b56\u7565\uff1asafe_only = \u53ea\u9ed8\u8ba4\u7eb3\u5165\u4f4e\u98ce\u9669\u8bb0\u5f55\uff1breview_all = \u5168\u90e8\u9700\u4eba\u5de5\u786e\u8ba4\uff1binclude_all_except_high = \u9664\u9ad8\u98ce\u9669\u5916\u9ed8\u8ba4\u7eb3\u5165\u3002
  "default_include_policy": "safe_only"
}
''')


def strip_jsonc_comments(text: str) -> str:
    out = []
    in_string = False
    escaped = False
    i = 0
    while i < len(text):
        ch = text[i]
        nxt = text[i + 1] if i + 1 < len(text) else ""
        if in_string:
            out.append(ch)
            if escaped:
                escaped = False
            elif ch == "\\":
                escaped = True
            elif ch == '"':
                in_string = False
            i += 1
            continue
        if ch == '"':
            in_string = True
            out.append(ch)
            i += 1
            continue
        if ch == "/" and nxt == "/":
            while i < len(text) and text[i] not in "\r\n":
                i += 1
            continue
        if ch == "/" and nxt == "*":
            i += 2
            while i + 1 < len(text) and not (text[i] == "*" and text[i + 1] == "/"):
                i += 1
            i += 2
            continue
        out.append(ch)
        i += 1
    return "".join(out)


def ensure_privacy_config(path: Path) -> None:
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(default_privacy_config_text(), encoding="utf-8")


def load_privacy_config(raw_path: str | None) -> tuple[dict[str, Any], Path]:
    path = Path(raw_path).expanduser().resolve() if raw_path else default_config_path()
    ensure_privacy_config(path)
    try:
        config = json.loads(strip_jsonc_comments(path.read_text(encoding="utf-8-sig")))
    except Exception as exc:
        raise SystemExit(f"privacy config parse failed: {path}: {exc}") from exc
    defaults = {"privacy_blacklist": [], "project_whitelist": [], "project_blacklist": [], "exclude_tools": [], "exclude_time_ranges": [], "default_include_policy": "safe_only"}
    defaults.update(config if isinstance(config, dict) else {})
    return defaults, path


def hhmm(value: str) -> int | None:
    try:
        hour, minute = value.split(":", 1)
        return int(hour) * 60 + int(minute)
    except Exception:
        return None


def overlaps_excluded_time(time_text: str, ranges: list[dict[str, Any]]) -> str | None:
    if not time_text or "-" not in time_text:
        return None
    start_text, end_text = time_text.split("-", 1)
    start = hhmm(start_text.strip())
    end = hhmm(end_text.strip())
    if start is None or end is None:
        return None
    for item in ranges or []:
        ex_start = hhmm(str(item.get("start", "")))
        ex_end = hhmm(str(item.get("end", "")))
        if ex_start is None or ex_end is None:
            continue
        if start < ex_end and end > ex_start:
            return str(item.get("reason") or f"{item.get('start')}-{item.get('end')}")
    return None


def contains_any(text: str, words: list[Any]) -> list[str]:
    lower = text.lower()
    return [str(word) for word in words or [] if str(word).lower() in lower]


def time_range_text(row: dict[str, Any]) -> str:
    start_time = str(row.get("start", ""))[-8:-3]
    end_time = str(row.get("end", ""))[-8:-3]
    return f"{start_time}-{end_time}" if start_time or end_time else ""


def evidence_file_name(value: Any) -> str:
    try:
        return Path(str(value)).name
    except Exception:
        return str(value or "")


def classify_privacy(row: dict[str, Any], config: dict[str, Any]) -> dict[str, Any]:
    text = " ".join(str(row.get(key) or "") for key in ["request", "result", "cwd", "evidence", "tool"])
    hits = contains_any(text, config.get("privacy_blacklist") or [])
    if hits:
        return {"risk": cn(r"\u9ad8"), "reason": cn(r"\u547d\u4e2d\u9690\u79c1\u8bcd\uff1a") + cn(r"\u3001").join(hits), "include": False, "suggestion": cn(r"\u5efa\u8bae\u6392\u9664")}
    tool_hits = contains_any(str(row.get("tool") or ""), config.get("exclude_tools") or [])
    if tool_hits:
        return {"risk": cn(r"\u9ad8"), "reason": cn(r"\u547d\u4e2d\u6392\u9664 AI \u5de5\u5177"), "include": False, "suggestion": cn(r"\u5efa\u8bae\u6392\u9664")}
    range_reason = overlaps_excluded_time(time_range_text(row), config.get("exclude_time_ranges") or [])
    if range_reason:
        return {"risk": cn(r"\u9ad8"), "reason": cn(r"\u547d\u4e2d\u6392\u9664\u65f6\u95f4\u6bb5\uff1a") + range_reason, "include": False, "suggestion": cn(r"\u5efa\u8bae\u6392\u9664")}
    project_black_hits = contains_any(text, config.get("project_blacklist") or [])
    if project_black_hits:
        return {"risk": cn(r"\u9ad8"), "reason": cn(r"\u547d\u4e2d\u9879\u76ee\u9ed1\u540d\u5355\uff1a") + cn(r"\u3001").join(project_black_hits), "include": False, "suggestion": cn(r"\u5efa\u8bae\u6392\u9664")}
    whitelist = config.get("project_whitelist") or []
    if whitelist and not contains_any(text, whitelist):
        return {"risk": cn(r"\u4e2d"), "reason": cn(r"\u672a\u547d\u4e2d\u9879\u76ee\u767d\u540d\u5355"), "include": False, "suggestion": cn(r"\u9700\u786e\u8ba4")}
    if not row.get("cwd") and not row.get("evidence"):
        return {"risk": cn(r"\u4e2d"), "reason": cn(r"\u7f3a\u5c11\u9879\u76ee\u6216\u8bc1\u636e\u5173\u8054"), "include": False, "suggestion": cn(r"\u9700\u786e\u8ba4")}
    policy = str(config.get("default_include_policy") or "safe_only")
    include = policy != "review_all"
    return {"risk": cn(r"\u4f4e"), "reason": cn(r"\u672a\u547d\u4e2d\u98ce\u9669\u89c4\u5219"), "include": include, "suggestion": cn(r"\u5efa\u8bae\u7eb3\u5165") if include else cn(r"\u9700\u786e\u8ba4")}



def apply_cli_filters(config: dict[str, Any], args: argparse.Namespace) -> dict[str, Any]:
    merged = dict(config)
    if getattr(args, "policy", None):
        merged["default_include_policy"] = args.policy
    if getattr(args, "include_keyword", None):
        merged["project_whitelist"] = [*list(merged.get("project_whitelist") or []), *args.include_keyword]
    if getattr(args, "exclude_keyword", None):
        merged["privacy_blacklist"] = [*list(merged.get("privacy_blacklist") or []), *args.exclude_keyword]
    if getattr(args, "exclude_tool", None):
        merged["exclude_tools"] = [*list(merged.get("exclude_tools") or []), *args.exclude_tool]
    if getattr(args, "exclude_time", None):
        ranges = list(merged.get("exclude_time_ranges") or [])
        for raw in args.exclude_time:
            if "-" not in raw:
                continue
            start, end = raw.split("-", 1)
            ranges.append({"start": start.strip(), "end": end.strip(), "reason": cn(r"\u4e34\u65f6\u5bf9\u8bdd\u8fc7\u6ee4")})
        merged["exclude_time_ranges"] = ranges
    return merged


def preview_summary(conversations: list[dict[str, Any]], config: dict[str, Any]) -> dict[str, Any]:
    summary = {"total": 0, "include": 0, "exclude": 0, "low": 0, "medium": 0, "high": 0, "tools": {}, "projects": {}}
    for row in conversations:
        verdict = classify_privacy(row, config)
        summary["total"] += 1
        summary["include" if verdict["include"] else "exclude"] += 1
        risk = str(verdict["risk"])
        if risk == cn(r"\u4f4e"):
            summary["low"] += 1
        elif risk == cn(r"\u4e2d"):
            summary["medium"] += 1
        elif risk == cn(r"\u9ad8"):
            summary["high"] += 1
        tool = str(row.get("tool") or "")
        summary["tools"][tool] = summary["tools"].get(tool, 0) + 1
        project = Path(str(row.get("cwd") or "")).name or cn(r"\u672a\u8bc6\u522b")
        summary["projects"][project] = summary["projects"].get(project, 0) + 1
    return summary


def write_state(path: Path, state: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def read_state(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        obj = json.loads(path.read_text(encoding="utf-8-sig"))
        return obj if isinstance(obj, dict) else {}
    except Exception:
        return {}

def write_preview_excel(conversations: list[dict[str, Any]], config: dict[str, Any], config_path: Path, output_path: Path, day: dt.date) -> dict[str, Any]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SystemExit(cn(r"\u672a\u5b89\u88c5 openpyxl\uff0c\u65e0\u6cd5\u5199\u5165 .xlsx\u3002\u8bf7\u5148\u6267\u884c\uff1apython -m pip install openpyxl")) from exc

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = cn(r"\u5019\u9009\u7d20\u6750\u6e05\u5355")
    headers = [
        cn(r"\u662f\u5426\u7eb3\u5165"), cn(r"\u5e8f\u53f7"), cn(r"\u9690\u79c1\u98ce\u9669"), cn(r"\u98ce\u9669\u539f\u56e0"), cn(r"\u6765\u6e90\u5de5\u5177"), cn(r"\u65e5\u671f"), cn(r"\u65f6\u95f4\u6bb5"), cn(r"\u7528\u6237\u8bf7\u6c42\u6458\u8981"), cn(r"AI\u5904\u7406\u6458\u8981"), cn(r"\u6d89\u53ca\u9879\u76ee"), cn(r"\u8bc1\u636e"), cn(r"\u6392\u9664\u5efa\u8bae"), cn(r"AI\u6d3b\u8dc3\u8017\u65f6(\u5206\u949f)"), cn(r"\u4f30\u7b97\u771f\u5b9e\u5de5\u4f5c\u8017\u65f6(\u5206\u949f)")
    ]
    ws.append(headers)
    for idx, row in enumerate(conversations, 1):
        verdict = classify_privacy(row, config)
        ws.append([
            cn(r"\u662f") if verdict["include"] else cn(r"\u5426"),
            idx,
            verdict["risk"],
            verdict["reason"],
            row.get("tool", ""),
            day.isoformat(),
            time_range_text(row),
            row.get("request", ""),
            row.get("result", ""),
            Path(str(row.get("cwd") or "")).name,
            evidence_file_name(row.get("evidence", "")),
            verdict["suggestion"],
            row.get("ai_active_minutes", ""),
            row.get("estimated_work_minutes", ""),
        ])

    rules = wb.create_sheet(cn(r"\u9690\u79c1\u89c4\u5219\u8bf4\u660e"))
    rules.append([cn(r"\u914d\u7f6e\u9879"), cn(r"\u5f53\u524d\u503c"), cn(r"\u8bf4\u660e")])
    rules.append(["privacy_blacklist", cn(r"\u3001").join(str(x) for x in config.get("privacy_blacklist", [])), cn(r"\u547d\u4e2d\u540e\u9ed8\u8ba4\u9ad8\u98ce\u9669\u5e76\u6392\u9664")])
    rules.append(["project_whitelist", cn(r"\u3001").join(str(x) for x in config.get("project_whitelist", [])), cn(r"\u975e\u7a7a\u65f6\u53ea\u9ed8\u8ba4\u7eb3\u5165\u547d\u4e2d\u9879\u76ee")])
    rules.append(["project_blacklist", cn(r"\u3001").join(str(x) for x in config.get("project_blacklist", [])), cn(r"\u547d\u4e2d\u540e\u9ed8\u8ba4\u6392\u9664")])
    rules.append(["exclude_tools", cn(r"\u3001").join(str(x) for x in config.get("exclude_tools", [])), cn(r"\u6392\u9664\u6307\u5b9a AI \u5de5\u5177\u6765\u6e90")])
    rules.append(["exclude_time_ranges", json.dumps(config.get("exclude_time_ranges", []), ensure_ascii=False), cn(r"\u6392\u9664\u6307\u5b9a\u65f6\u95f4\u6bb5")])
    rules.append(["default_include_policy", config.get("default_include_policy", "safe_only"), cn(r"safe_only/review_all/include_all_except_high")])
    rules.append([cn(r"\u672c\u5730\u914d\u7f6e\u6587\u4ef6"), str(config_path), cn(r"\u53ef\u76f4\u63a5\u4fee\u6539\uff0c\u4e0d\u63d0\u4ea4 Git")])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for sheet_obj in wb.worksheets:
        sheet_obj.freeze_panes = "A2"
        sheet_obj.auto_filter.ref = sheet_obj.dimensions
        for cell in sheet_obj[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in sheet_obj.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, column_cells in enumerate(sheet_obj.columns, 1):
            max_len = max((len(str(cell.value or "")) for cell in list(column_cells)[:120]), default=0)
            sheet_obj.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 70)
    wb.save(output_path)
    return {"path": str(output_path), "rows": len(conversations), "config": str(config_path)}


def read_confirmed_preview(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(cn(r"\u672a\u5b89\u88c5 openpyxl\uff0c\u65e0\u6cd5\u8bfb\u53d6 .xlsx\u3002\u8bf7\u5148\u6267\u884c\uff1apython -m pip install openpyxl")) from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = cn(r"\u5019\u9009\u7d20\u6750\u6e05\u5355")
    if sheet_name not in wb.sheetnames:
        raise SystemExit(cn(r"preview Excel \u7f3a\u5c11\u5de5\u4f5c\u8868\uff1a\u5019\u9009\u7d20\u6750\u6e05\u5355"))
    ws = wb[sheet_name]
    cols = {str(ws.cell(1, col).value or ""): col for col in range(1, ws.max_column + 1)}
    yes_values = {cn(r"\u662f"), "Y", "y", "YES", "yes", "1", "true", "TRUE"}
    required = [cn(r"\u662f\u5426\u7eb3\u5165"), cn(r"\u6765\u6e90\u5de5\u5177"), cn(r"\u65e5\u671f"), cn(r"\u65f6\u95f4\u6bb5"), cn(r"\u7528\u6237\u8bf7\u6c42\u6458\u8981"), cn(r"AI\u5904\u7406\u6458\u8981")]
    missing = [name for name in required if name not in cols]
    if missing:
        raise SystemExit(cn(r"preview Excel \u7f3a\u5c11\u5217\uff1a") + ", ".join(missing))
    rows: list[dict[str, Any]] = []
    for row_idx in range(2, ws.max_row + 1):
        include = str(ws.cell(row_idx, cols[cn(r"\u662f\u5426\u7eb3\u5165")]).value or "").strip()
        if include not in yes_values:
            continue
        date_text = str(ws.cell(row_idx, cols[cn(r"\u65e5\u671f")]).value or "")
        time_text = str(ws.cell(row_idx, cols[cn(r"\u65f6\u95f4\u6bb5")]).value or "")
        start_time, end_time = (time_text.split("-", 1) + [""])[:2] if "-" in time_text else ("", "")
        def cell(name: str) -> Any:
            return ws.cell(row_idx, cols[name]).value if name in cols else ""
        rows.append({
            "tool": cell(cn(r"\u6765\u6e90\u5de5\u5177")),
            "start": f"{date_text} {start_time}:00" if start_time else date_text,
            "end": f"{date_text} {end_time}:00" if end_time else date_text,
            "request": cell(cn(r"\u7528\u6237\u8bf7\u6c42\u6458\u8981")),
            "result": cell(cn(r"AI\u5904\u7406\u6458\u8981")),
            "cwd": cell(cn(r"\u6d89\u53ca\u9879\u76ee")),
            "evidence": cell(cn(r"\u8bc1\u636e")),
            "ai_active_minutes": float(cell(cn(r"AI\u6d3b\u8dc3\u8017\u65f6(\u5206\u949f)")) or 0),
            "estimated_work_minutes": float(cell(cn(r"\u4f30\u7b97\u771f\u5b9e\u5de5\u4f5c\u8017\u65f6(\u5206\u949f)")) or 0),
        })
    return rows


def task_type_label(row: dict[str, Any]) -> str:
    text = f"{row.get('request', '')} {row.get('result', '')}".lower()
    if any(k in text for k in ["bug", "fix", "commit", "mapper", cn(r"\u4fee\u590d"), cn(r"\u63d0\u4ea4"), cn(r"\u53e3\u5f84")]):
        return cn(r"\u4ee3\u7801\u5f00\u53d1 / Bug \u4fee\u590d")
    if any(k in text for k in ["sql", "postgres", "pg", cn(r"\u6570\u636e\u5e93"), cn(r"\u67e5\u8be2"), cn(r"\u9a8c\u8bc1")]):
        return cn(r"SQL/\u6570\u636e\u9a8c\u8bc1")
    if any(k in text for k in ["skill", "script", "tool", cn(r"\u811a\u672c"), cn(r"\u5de5\u5177"), cn(r"\u81ea\u52a8\u5316")]):
        return cn(r"\u5de5\u5177/skill \u5efa\u8bbe")
    if any(k in text for k in ["sheet", "excel", "doc", "lark", cn(r"\u98de\u4e66"), cn(r"\u8868\u683c"), cn(r"\u6587\u6863")]):
        return cn(r"\u6587\u6863/\u8868\u683c\u5206\u6790")
    if any(k in text for k in [cn(r"\u6392\u67e5"), cn(r"\u5206\u6790"), cn(r"\u94fe\u8def")]):
        return cn(r"\u95ee\u9898\u6392\u67e5")
    return cn(r"\u65e5\u5e38\u534f\u4f5c")


def public_task_phrase(row: dict[str, Any]) -> str:
    request = strip_public_noise(row.get("request", ""))
    result = strip_public_noise(row.get("result", ""))
    text = f"{request} {result}".lower()
    if any(k in text for k in ["skill", "script", "tool", cn(r"\u811a\u672c"), cn(r"\u5de5\u5177"), cn(r"\u81ea\u52a8\u5316")]):
        return cn(r"\u4f18\u5316\u81ea\u52a8\u5316\u5de5\u5177\u548c Skill \u5de5\u4f5c\u6d41\uff0c\u63d0\u5347\u8d44\u6599\u8bfb\u53d6\u3001\u5904\u7406\u548c\u65e5\u62a5\u5f52\u6863\u6548\u7387")
    elif any(k in text for k in ["bug", "fix", cn(r"\u4fee\u590d"), cn(r"\u9519\u4f4d"), cn(r"\u5f02\u5e38")]):
        return cn(r"\u5b8c\u6210\u4e1a\u52a1\u6570\u636e\u5c55\u793a\u548c\u5bfc\u51fa\u95ee\u9898\u4fee\u590d\uff0c\u4fdd\u969c\u524d\u540e\u53f0\u6570\u636e\u4e00\u81f4\u6027")
    elif any(k in text for k in ["sql", "postgres", "pg", cn(r"\u6570\u636e\u5e93"), cn(r"\u67e5\u8be2"), cn(r"\u9a8c\u8bc1")]):
        return cn(r"\u5b8c\u6210\u6570\u636e\u6838\u5bf9\u3001SQL \u9a8c\u8bc1\u548c\u5f02\u5e38\u53e3\u5f84\u5206\u6790\uff0c\u4e3a\u540e\u7eed\u4fee\u590d\u63d0\u4f9b\u4f9d\u636e")
    elif any(k in text for k in ["sheet", "excel", "doc", "lark", cn(r"\u98de\u4e66"), cn(r"\u8868\u683c"), cn(r"\u6587\u6863")]):
        return cn(r"\u68b3\u7406\u6587\u6863\u548c\u8868\u683c\u8d44\u6599\uff0c\u5b8c\u6210\u5185\u5bb9\u6838\u5bf9\u3001\u8bfb\u53d6\u548c\u53ef\u7528\u6027\u5224\u65ad")
    elif any(k in text for k in [cn(r"\u6392\u67e5"), cn(r"\u5206\u6790"), cn(r"\u94fe\u8def")]):
        return cn(r"\u68b3\u7406\u4e1a\u52a1\u548c\u7cfb\u7edf\u5904\u7406\u94fe\u8def\uff0c\u660e\u786e\u63a5\u53e3\u3001\u6570\u636e\u6e90\u548c\u6392\u67e5\u65b9\u5411")
    return cn(r"\u63a8\u8fdb\u65e5\u5e38\u534f\u4f5c\u4e8b\u9879\u6574\u7406\u548c\u4ea4\u4ed8\u8ddf\u8fdb")


def public_project_label(row: dict[str, Any]) -> str:
    project = Path(str(row.get("cwd") or "")).name.strip()
    return project or cn(r"\u672a\u5f52\u5c5e\u9879\u76ee")


def public_feature_label(row: dict[str, Any]) -> str:
    text = f"{row.get('request', '')} {row.get('result', '')}".lower()
    if any(k in text for k in ["ai-worklog", "worklog", "preview", "codex", "claude", "session", cn(r"\u62a5\u5de5"), cn(r"\u65e5\u62a5"), cn(r"AI \u534f\u4f5c"), cn(r"\u9690\u79c1")]):
        return cn(r"AI \u5de5\u4f5c\u65e5\u62a5 / \u62a5\u5de5\u5de5\u5177")
    if any(k in text for k in ["install-codexskill", "install.ps1", "force", "backup", cn(r"\u5b89\u88c5"), cn(r"\u5907\u4efd"), cn(r"\u914d\u7f6e\u6587\u4ef6")]):
        return cn(r"Skill \u5b89\u88c5\u4e0e\u672c\u5730\u914d\u7f6e\u4fdd\u7559")
    if any(k in text for k in ["lark", cn(r"\u98de\u4e66"), "im:", cn(r"\u804a\u5929\u8bb0\u5f55")]):
        return cn(r"\u98de\u4e66\u6388\u6743\u4e0e\u6570\u636e\u8bfb\u53d6")
    if any(k in text for k in ["selection-detail-task-export", "export-all-task", cn(r"\u63a8\u9001\u65e5\u5fd7"), cn(r"\u5bfc\u51fa")]):
        return cn(r"OTB \u63a8\u9001\u65e5\u5fd7 / \u5bfc\u51fa\u80fd\u529b")
    if any(k in text for k in ["channel-otb", "channel_otb", cn(r"\u6e20\u9053otb"), "store-type", "region", cn(r"\u6e29\u533a"), cn(r"\u7ba1\u7406\u533a\u57df"), cn(r"\u95e8\u5e97")]):
        return cn(r"\u6e20\u9053 OTB \u6570\u636e\u53e3\u5f84")
    if any(k in text for k in ["signature_direct_sales", cn(r"\u7b7e\u6279\u8868"), cn(r"\u4e70\u8d27\u5e97\u94fa\u6570")]):
        return cn(r"OTB \u7b7e\u6279\u8868\u6570\u636e\u53e3\u5f84")
    if any(k in text for k in ["sql", "uat", "dev", "dazy", cn(r"\u6570\u636e\u5e93"), cn(r"\u8bc1\u660e"), cn(r"\u6838\u5bf9")]):
        return cn(r"\u6570\u636e\u6838\u5bf9\u4e0e SQL \u8bc1\u660e")
    return task_type_label(row)


def public_delivery_point(row: dict[str, Any]) -> str:
    text = f"{row.get('request', '')} {row.get('result', '')}".lower()
    if any(k in text for k in ["ai-worklog", "worklog", "codex", "claude", "session", cn(r"\u62a5\u5de5"), cn(r"\u65e5\u62a5"), cn(r"AI \u534f\u4f5c")]):
        return cn(r"\u5b8c\u6210 AI \u534f\u4f5c\u8bb0\u5f55\u7684\u9884\u89c8\u786e\u8ba4\u3001\u9690\u79c1\u8fc7\u6ee4\u3001Excel \u53f0\u8d26\u548c\u62a5\u5de5\u6458\u8981\u6d41\u7a0b\u68b3\u7406\u3002")
    if any(k in text for k in ["install-codexskill", "install.ps1", "force", "backup", cn(r"\u5907\u4efd"), cn(r"\u914d\u7f6e\u6587\u4ef6")]):
        return cn(r"\u4f18\u5316 skill \u8986\u76d6\u5b89\u88c5\u903b\u8f91\uff0c\u652f\u6301\u65e7\u7248\u5b8c\u6574\u5907\u4efd\u3001\u672c\u5730\u914d\u7f6e\u7ed3\u6784\u6821\u9a8c\u548c\u540c\u7ed3\u6784\u81ea\u52a8\u4fdd\u7559\u3002")
    if any(k in text for k in ["lark", cn(r"\u98de\u4e66"), "im:"]):
        return cn(r"\u68b3\u7406\u98de\u4e66\u804a\u5929\u8bb0\u5f55\u8bfb\u53d6\u53ef\u884c\u6027\u548c lark-cli \u6743\u9650\u8981\u6c42\uff0c\u660e\u786e\u5f53\u524d\u6388\u6743\u7f3a\u53e3\u3002")
    if any(k in text for k in ["selection-detail-task-export", "export-all-task", cn(r"\u63a8\u9001\u65e5\u5fd7"), cn(r"\u5bfc\u51fa")]):
        return cn(r"\u5206\u6790\u63a8\u9001\u65e5\u5fd7\u5217\u8868\u94fe\u8def\uff0c\u8865\u5145\u5168\u91cf\u5bfc\u51fa\u63a5\u53e3\u4e0e\u201c\u63a8\u9001\u65e5\u5fd7\u5bfc\u51fa\u201d\u9875\u9762\u914d\u7f6e SQL\u3002")
    if any(k in text for k in ["channel-otb", "channel_otb", cn(r"\u6e20\u9053otb"), "store-type", "region", cn(r"\u6e29\u533a"), cn(r"\u7ba1\u7406\u533a\u57df")]):
        return cn(r"\u5bf9\u9f50\u6e20\u9053 OTB \u95e8\u5e97\u6709\u6548\u6027\u3001\u95ed\u5e97/\u96f6\u552e\u76ee\u6807\u3001\u5fae\u5546\u57ce/\u603b\u90e8\u4ed3\u5e93\u548c zero-row \u8fc7\u6ee4\u53e3\u5f84\u3002")
    if any(k in text for k in ["signature_direct_sales", cn(r"\u7b7e\u6279\u8868"), cn(r"\u4e70\u8d27\u5e97\u94fa\u6570")]):
        return cn(r"\u5206\u6790 OTB \u7b7e\u6279\u8868\u4e0e\u6e20\u9053 OTB \u9875\u9762\u5e97\u94fa\u6570\u5dee\u5f02\uff0c\u660e\u786e\u6765\u6e90\u8868\u548c\u516c\u5f0f\u53e3\u5f84\u3002")
    if any(k in text for k in ["sql", "uat", "dev", "dazy", cn(r"\u6570\u636e\u5e93"), cn(r"\u8bc1\u660e"), cn(r"\u6838\u5bf9")]):
        return cn(r"\u63d0\u4f9b\u5dee\u5f02\u6570\u636e\u6838\u5bf9 SQL\uff0c\u652f\u6301\u6309\u95e8\u5e97\u7f16\u7801\u7cbe\u786e\u5b9a\u4f4d\u53c2\u6570\u4e0d\u4e00\u81f4\u7684\u660e\u7ec6\u884c\u3002")
    return public_task_phrase(row)


def strip_public_noise(text: Any) -> str:
    value = shorten(text, 120)
    value = re.sub(r"https?://\S+", cn(r"\u76f8\u5173\u94fe\u63a5"), value)
    value = re.sub(r"[A-Za-z]:\\[^\s]+", cn(r"\u672c\u5730\u8def\u5f84"), value)
    return value


def count_git_commits(git_rows: list[dict[str, Any]]) -> int:
    total = 0
    for row in git_rows:
        commits = str(row.get("commits") or "").strip()
        if not commits or commits.startswith("git_error"):
            continue
        total += len([line for line in commits.splitlines() if line.strip()])
    return total


def build_public_report_text(date_value: str, conversations: list[dict[str, Any]], git_rows: list[dict[str, Any]]) -> str:
    if not conversations:
        return date_value + cn(r"\u672a\u53d1\u73b0\u53ef\u7eb3\u5165\u62a5\u5de5\u7684 AI \u534f\u4f5c\u8bb0\u5f55\u3002")
    categories: dict[str, int] = {}
    projects: list[str] = []
    for row in conversations:
        label = task_type_label(row)
        categories[label] = categories.get(label, 0) + 1
        project = Path(str(row.get("cwd") or "")).name.strip()
        if project and project not in projects:
            projects.append(project)
    top_phrases = []
    seen = set()
    for row in conversations:
        phrase = public_task_phrase(row)
        key = phrase[:50]
        if key in seen:
            continue
        seen.add(key)
        top_phrases.append(phrase)
        if len(top_phrases) >= 6:
            break
    total_work = round(sum(float(row.get("estimated_work_minutes") or 0) for row in conversations), 1)
    hours = round(total_work / 60, 1)
    category_text = cn(r"\u3001").join(f"{name}{count}" + cn(r"\u9879") for name, count in sorted(categories.items(), key=lambda item: item[1], reverse=True)[:4])
    project_text = cn(r"\u3001").join(projects[:5]) if projects else cn(r"\u591a\u4e2a\u9879\u76ee")
    phrase_text = cn(r"\uff1b").join(top_phrases)
    commit_count = count_git_commits(git_rows)
    commit_text = ""
    if commit_count:
        commit_text = cn(r"\u5e76\u5b8c\u6210 ") + str(commit_count) + cn(r" \u4e2a Git \u63d0\u4ea4\u6216\u4ee3\u7801\u8bc1\u636e\u5f52\u6863\uff0c\u4fbf\u4e8e\u540e\u7eed\u8ffd\u6eaf\u3002")
    return (
        date_value
        + cn(r"\u5b8c\u6210 ")
        + project_text
        + cn(r" \u76f8\u5173\u5de5\u4f5c\uff0c\u5171\u5f52\u6863 ")
        + str(len(conversations))
        + cn(r" \u6761\u7ecf\u786e\u8ba4\u7684\u5de5\u4f5c\u7d20\u6750\uff0c\u8986\u76d6")
        + category_text
        + cn(r"\u3002\u4e3b\u8981\u4ea4\u4ed8\u5305\u62ec\uff1a")
        + phrase_text
        + cn(r"\u3002")
        + commit_text
        + cn(r"\u5efa\u8bae\u62a5\u5de5\u5de5\u65f6\u7ea6 ")
        + str(hours)
        + cn(r" \u5c0f\u65f6\uff0c\u5177\u4f53\u53ef\u6309\u5355\u4f4d\u62a5\u5de5\u53e3\u5f84\u8c03\u6574\u3002")
    )


def build_public_report_rows(date_value: str, conversations: list[dict[str, Any]], git_rows: list[dict[str, Any]]) -> list[list[Any]]:
    if not conversations:
        return [[cn(r"\u672a\u53d1\u73b0\u53ef\u7eb3\u5165\u62a5\u5de5\u7684 AI \u534f\u4f5c\u8bb0\u5f55"), "", "", 0, "", ""]]

    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    project_order: list[str] = []
    for row in conversations:
        project = public_project_label(row)
        feature = public_feature_label(row)
        if feature in {cn(r"\u65e5\u5e38\u534f\u4f5c"), cn(r"\u4ee3\u7801\u5f00\u53d1 / Bug \u4fee\u590d")} and float(row.get("estimated_work_minutes") or 0) < 1:
            feature = cn(r"AI \u5de5\u4f5c\u65e5\u62a5 / \u62a5\u5de5\u5de5\u5177")
        if project not in project_order:
            project_order.append(project)
        key = (project, feature)
        item = grouped.setdefault(key, {"count": 0, "minutes": 0.0, "points": [], "evidence": set()})
        item["count"] += 1
        item["minutes"] += float(row.get("estimated_work_minutes") or 0)
        point = public_delivery_point(row)
        if point and point not in item["points"]:
            item["points"].append(point)
        evidence = evidence_file_name(row.get("evidence", ""))
        if evidence:
            item["evidence"].add(evidence)

    commit_count = count_git_commits(git_rows)
    rows: list[list[Any]] = []
    index = 1
    for project in project_order:
        feature_items = [(feature, grouped[(project, feature)]) for feature in sorted([f for p, f in grouped if p == project], key=lambda f: grouped[(project, f)]["minutes"], reverse=True)]
        for feature, item in feature_items:
            points = item["points"][:3]
            content_lines = [f"{idx}. {point}" for idx, point in enumerate(points, 1)]
            if len(item["points"]) > 3:
                content_lines.append(cn(r"\u5176\u4ed6\u76f8\u5173\u6c9f\u901a\u3001\u9a8c\u8bc1\u548c\u8ddf\u8fdb\u4e8b\u9879\u5df2\u5408\u5e76\u5f52\u6863\u3002"))
            evidence_parts = [cn(r"\u5f52\u5e76 AI \u534f\u4f5c\u8bb0\u5f55 ") + str(item["count"]) + cn(r" \u6761")]
            if item["evidence"]:
                evidence_parts.append(cn(r"\u4f1a\u8bdd\u8bc1\u636e ") + str(len(item["evidence"])) + cn(r" \u4efd"))
            if commit_count and project in {Path(str(row.get("repo") or "")).name for row in git_rows}:
                evidence_parts.append(cn(r"\u542b Git \u63d0\u4ea4\u8bc1\u636e"))
            rows.append([
                index,
                project,
                feature,
                "\n".join(content_lines),
                item["count"],
                cn(r"\u7ea6 ") + str(round(item["minutes"] / 60, 1)) + cn(r" \u5c0f\u65f6"),
                cn(r"\uff1b").join(evidence_parts),
            ])
            index += 1
    return rows

def write_excel(data: dict[str, Any], output_path: Path, mode: str = "upsert") -> dict[str, Any]:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SystemExit(cn(r"\u672a\u5b89\u88c5 openpyxl\uff0c\u65e0\u6cd5\u5199\u5165 .xlsx\u3002\u8bf7\u5148\u6267\u884c\uff1apython -m pip install openpyxl")) from exc

    date_value = str(data["date"])
    report = data.get("report") or {}
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if output_path.exists():
        wb = load_workbook(output_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    def sheet(name: str, headers: list[str]):
        full_headers = [cn(r"\u65e5\u671f"), *headers]
        if name in wb.sheetnames:
            ws = wb[name]
            existing = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
            if existing == full_headers:
                pass
            elif existing and existing[0] == cn(r"\u65e5\u671f"):
                for header in full_headers:
                    if header not in existing:
                        ws.cell(row=1, column=ws.max_column + 1, value=header)
                        existing.append(header)
            else:
                ws.delete_rows(1, ws.max_row)
                ws.append(full_headers)
        else:
            ws = wb.create_sheet(name)
            ws.append(full_headers)
        if ws.max_row == 0:
            ws.append(full_headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        if mode == "upsert":
            for row_idx in range(ws.max_row, 1, -1):
                if str(ws.cell(row=row_idx, column=1).value or "") == date_value:
                    ws.delete_rows(row_idx, 1)
        return ws

    def append_rows(ws, rows: list[list[Any]]):
        for row in rows:
            ws.append([date_value, *[redact(cell) for cell in row]])

    def value(row: dict[str, Any], keys: list[str], default: str = "") -> Any:
        for key in keys:
            if key in row:
                return row.get(key)
        return default

    def fmt_minutes(raw: Any) -> str:
        text = str(raw or "").strip()
        if text:
            return text if not text.replace(".", "", 1).isdigit() else f"{text} " + cn(r"\u5206\u949f")
        try:
            minutes = float(raw or 0)
        except (TypeError, ValueError):
            minutes = 0.0
        return f"{round(minutes, 1)} " + cn(r"\u5206\u949f")

    def time_range(row: dict[str, Any]) -> str:
        start = str(row.get("start", ""))[-8:-3]
        end = str(row.get("end", ""))[-8:-3]
        return f"{start}-{end}" if start or end else ""

    def evidence_name(raw: Any) -> str:
        try:
            return Path(str(raw)).name
        except Exception:
            return str(raw or "")

    def evidence_strength(*parts: Any) -> str:
        text = " ".join(str(part or "") for part in parts).lower()
        if ("commit" in text or re.search(r"\b[0-9a-f]{7,40}\b", text)) and ("session" in text or "??" in text or "file" in text or "sql" in text):
            return cn(r"\u5f3a")
        if "commit" in text or re.search(r"\b[0-9a-f]{7,40}\b", text):
            return cn(r"\u5f3a")
        if any(k in text for k in ["session", "sql", "pg", "postgres", "??", "??", "????", "file"]):
            return cn(r"\u4e2d")
        return cn(r"\u5f31")

    conversations = data.get("conversations") or []
    git_rows = data.get("git") or []
    sources = data.get("sources") or []

    source_rows = report.get("sources") or []
    if source_rows and isinstance(source_rows[0], dict):
        source_rows = [[value(r, ["source", cn(r"\u6765\u6e90")]), value(r, ["path", "repo", cn(r"\u8def\u5f84/\u4ed3\u5e93")]), value(r, ["has_record", cn(r"\u662f\u5426\u6709\u8bb0\u5f55")]), value(r, ["note", cn(r"\u8bf4\u660e")])] for r in source_rows]
    elif not source_rows:
        source_rows = []
        for src in sources:
            name = src.get("source", "")
            label = "Claude Code" if name == "Claude" else name
            status = cn(r"\u6709") if src.get("status") == "found" else cn(r"\u65e0")
            source_rows.append([label, src.get("path", ""), status, str(src.get("count", 0))])
    ws = sheet(cn(r"\u6570\u636e\u6765\u6e90\u68c0\u67e5"), [cn(r"\u6765\u6e90"), cn(r"\u8def\u5f84/\u4ed3\u5e93"), cn(r"\u662f\u5426\u6709\u8bb0\u5f55"), cn(r"\u8bf4\u660e")])
    append_rows(ws, source_rows)

    conv_rows = report.get("conversation_summary") or []
    if conv_rows and isinstance(conv_rows[0], dict):
        conv_rows = [[value(r, ["index", cn(r"\u5e8f\u53f7")]), value(r, ["tool", cn(r"AI\u5de5\u5177")]), value(r, ["time", cn(r"\u65f6\u95f4")]), value(r, ["request", cn(r"\u7528\u6237\u8bf7\u6c42\u6458\u8981")]), value(r, ["result", cn(r"AI\u5904\u7406\u7ed3\u679c")]), value(r, ["ai_active", cn(r"AI\u6d3b\u8dc3\u8017\u65f6")]), value(r, ["evidence", cn(r"\u8bc1\u636e")])] for r in conv_rows]
    elif not conv_rows:
        conv_rows = [[idx, row.get("tool", ""), time_range(row), row.get("request", ""), row.get("result", ""), fmt_minutes(row.get("ai_active_minutes")), evidence_file_name(row.get("evidence", ""))] for idx, row in enumerate(conversations, 1)]
    ws = sheet(cn(r"AI \u5bf9\u8bdd\u660e\u7ec6"), [cn(r"\u5e8f\u53f7"), cn(r"AI\u5de5\u5177"), cn(r"\u65f6\u95f4"), cn(r"\u7528\u6237\u8bf7\u6c42\u6458\u8981"), cn(r"AI\u5904\u7406\u7ed3\u679c"), cn(r"AI\u6d3b\u8dc3\u8017\u65f6"), cn(r"\u8bc1\u636e")])
    append_rows(ws, conv_rows)

    task_rows = report.get("task_summary") or []
    if task_rows and isinstance(task_rows[0], dict):
        task_rows = [[value(r, ["index", cn(r"\u5e8f\u53f7")]), value(r, ["task", cn(r"\u4efb\u52a1")]), value(r, ["type", cn(r"\u7c7b\u578b")]), value(r, ["project", cn(r"\u6d89\u53ca\u9879\u76ee")]), value(r, ["time", cn(r"\u8d77\u6b62\u65f6\u95f4")]), value(r, ["ai_active", cn(r"AI\u6d3b\u8dc3\u8017\u65f6")]), value(r, ["estimated_work", cn(r"\u4f30\u7b97\u771f\u5b9e\u5de5\u4f5c\u8017\u65f6")]), value(r, ["delivery", cn(r"\u4ea4\u4ed8\u7ed3\u679c")]), value(r, ["status", cn(r"\u72b6\u6001")]), value(r, ["evidence_strength", cn(r"\u8bc1\u636e\u5f3a\u5ea6")]) or evidence_strength(value(r, ["task", cn(r"\u4efb\u52a1")]), value(r, ["delivery", cn(r"\u4ea4\u4ed8\u7ed3\u679c")]), value(r, ["status", cn(r"\u72b6\u6001")]))] for r in task_rows]
    elif not task_rows:
        task_rows = [[idx, row.get("request", ""), cn(r"\u672a\u5f52\u5e76"), Path(str(row.get("cwd") or "")).name, time_range(row), fmt_minutes(row.get("ai_active_minutes")), fmt_minutes(row.get("estimated_work_minutes")), row.get("result", ""), cn(r"\u5df2\u5206\u6790"), evidence_strength(row.get("request", ""), row.get("result", ""), row.get("evidence", ""))] for idx, row in enumerate(conversations, 1)]
    ws = sheet(cn(r"\u4efb\u52a1\u5f52\u5e76\u7edf\u8ba1"), [cn(r"\u5e8f\u53f7"), cn(r"\u4efb\u52a1"), cn(r"\u7c7b\u578b"), cn(r"\u6d89\u53ca\u9879\u76ee"), cn(r"\u8d77\u6b62\u65f6\u95f4"), cn(r"AI\u6d3b\u8dc3\u8017\u65f6"), cn(r"\u4f30\u7b97\u771f\u5b9e\u5de5\u4f5c\u8017\u65f6"), cn(r"\u4ea4\u4ed8\u7ed3\u679c"), cn(r"\u72b6\u6001"), cn(r"\u8bc1\u636e\u5f3a\u5ea6")])
    append_rows(ws, task_rows)

    evidence_rows = report.get("evidence") or []
    if evidence_rows and isinstance(evidence_rows[0], dict):
        evidence_rows = [[value(r, ["type", cn(r"\u7c7b\u578b")]), value(r, ["location", "project", cn(r"\u9879\u76ee/\u4f4d\u7f6e")]), value(r, ["evidence", cn(r"\u8bc1\u636e")]), value(r, ["note", cn(r"\u8bf4\u660e")])] for r in evidence_rows]
    elif not evidence_rows:
        evidence_rows = []
        for row in git_rows:
            repo = Path(str(row.get("repo") or "")).name
            commits = str(row.get("commits") or cn(r"\u65e0\u5f53\u5929 commit"))
            for line in commits.splitlines() or [commits]:
                evidence_rows.append(["Git commit", repo, line, cn(r"Git \u5f53\u5929\u63d0\u4ea4")])
    ws = sheet(cn(r"\u8bc1\u636e\u6c47\u603b"), [cn(r"\u7c7b\u578b"), cn(r"\u9879\u76ee/\u4f4d\u7f6e"), cn(r"\u8bc1\u636e"), cn(r"\u8bf4\u660e")])
    append_rows(ws, evidence_rows)

    time_rows = report.get("time_summary") or []
    if time_rows and isinstance(time_rows[0], dict):
        time_rows = [[value(r, ["category", cn(r"\u5206\u7c7b")]), value(r, ["ai_active", cn(r"AI\u6d3b\u8dc3\u8017\u65f6")]), value(r, ["estimated_work", cn(r"\u4f30\u7b97\u771f\u5b9e\u5de5\u4f5c\u8017\u65f6")]), value(r, ["note", cn(r"\u8bf4\u660e")])] for r in time_rows]
    elif not time_rows:
        total_ai = round(sum(float(row.get("ai_active_minutes") or 0) for row in conversations), 1)
        total_work = round(sum(float(row.get("estimated_work_minutes") or 0) for row in conversations), 1)
        time_rows = [[cn(r"\u5408\u8ba1"), fmt_minutes(total_ai), fmt_minutes(total_work), cn(r"\u81ea\u52a8\u4f30\u7b97")]]
    ws = sheet(cn(r"\u8017\u65f6\u6c47\u603b"), [cn(r"\u5206\u7c7b"), cn(r"AI\u6d3b\u8dc3\u8017\u65f6"), cn(r"\u4f30\u7b97\u771f\u5b9e\u5de5\u4f5c\u8017\u65f6"), cn(r"\u8bf4\u660e")])
    append_rows(ws, time_rows)

    completion_rows = report.get("completion_summary") or []
    if completion_rows and isinstance(completion_rows[0], dict):
        completion_rows = [[value(r, ["index", cn(r"\u5e8f\u53f7")]), value(r, ["item", cn(r"\u5b8c\u6210\u4e8b\u9879")])] for r in completion_rows]
    elif completion_rows:
        completion_rows = [[idx, item] for idx, item in enumerate(completion_rows, 1)]
    else:
        completion_rows = [[idx, row.get("result") or row.get("request") or ""] for idx, row in enumerate(conversations[:20], 1)]
    ws = sheet(cn(r"\u4eca\u65e5\u5b8c\u6210\u4e8b\u9879\u6458\u8981"), [cn(r"\u5e8f\u53f7"), cn(r"\u5b8c\u6210\u4e8b\u9879")])
    append_rows(ws, completion_rows)

    public_rows = report.get("public_report") or report.get("report_rows") or []
    if public_rows and isinstance(public_rows[0], dict):
        public_rows = [[value(r, ["index", cn(r"\u5e8f\u53f7")]), value(r, ["project", cn(r"\u9879\u76ee")]), value(r, ["feature", cn(r"\u529f\u80fd/\u4e8b\u9879")]), value(r, ["content", cn(r"\u62a5\u5de5\u5185\u5bb9")]), value(r, ["count", cn(r"\u5f52\u5e76\u8bb0\u5f55\u6570")]), value(r, ["estimated_work", cn(r"\u5efa\u8bae\u62a5\u5de5\u5de5\u65f6")]), value(r, ["evidence", cn(r"\u8bc1\u636e\u6458\u8981")])] for r in public_rows]
    elif public_rows:
        normalized_rows = []
        for idx, item in enumerate(public_rows, 1):
            if isinstance(item, (list, tuple)):
                normalized_rows.append(list(item))
            else:
                normalized_rows.append([idx, "", "", item, "", "", ""])
        public_rows = normalized_rows
    else:
        report_text = report.get("report_text") or ""
        if report_text:
            public_rows = [[1, "", cn(r"\u81ea\u5b9a\u4e49\u62a5\u5de5\u5185\u5bb9"), report_text, "", "", ""]]
        else:
            public_rows = build_public_report_rows(date_value, conversations, git_rows)
    ws = sheet(cn(r"\u53ef\u76f4\u63a5\u62a5\u5de5\u7248\u672c"), [cn(r"\u5e8f\u53f7"), cn(r"\u9879\u76ee"), cn(r"\u529f\u80fd/\u4e8b\u9879"), cn(r"\u62a5\u5de5\u5185\u5bb9"), cn(r"\u5f52\u5e76\u8bb0\u5f55\u6570"), cn(r"\u5efa\u8bae\u62a5\u5de5\u5de5\u65f6"), cn(r"\u8bc1\u636e\u6458\u8981")])
    append_rows(ws, public_rows)

    def parse_minutes(raw: Any) -> float:
        text = str(raw or "").strip()
        if not text:
            return 0.0
        match = re.search(r"\d+(?:\.\d+)?", text)
        if not match:
            return 0.0
        value = float(match.group(0))
        if cn(r"\u5c0f\u65f6") in text:
            value *= 60
        return value

    def format_duration(minutes: float) -> str:
        if minutes >= 60:
            return cn(r"\u7ea6 ") + str(round(minutes / 60, 1)) + " " + cn(r"\u5c0f\u65f6")
        return cn(r"\u7ea6 ") + str(round(minutes, 1)).rstrip("0").rstrip(".") + " " + cn(r"\u5206\u949f")

    def headers(ws):
        return {str(ws.cell(1, col).value or ""): col for col in range(1, ws.max_column + 1)}

    def records(sheet_name: str) -> list[dict[str, Any]]:
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        cols = headers(ws)
        rows = []
        for row_idx in range(2, ws.max_row + 1):
            row = {name: ws.cell(row_idx, col).value for name, col in cols.items()}
            if row.get(cn(r"\u65e5\u671f")):
                rows.append(row)
        return rows

    def reset_summary_sheet(name: str, sheet_headers: list[str], index: int):
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name, index)
        ws.append(sheet_headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        return ws

    task_records = records(cn(r"\u4efb\u52a1\u5f52\u5e76\u7edf\u8ba1"))
    time_records = records(cn(r"\u8017\u65f6\u6c47\u603b"))
    evidence_records = records(cn(r"\u8bc1\u636e\u6c47\u603b"))

    dates = sorted({str(row.get(cn(r"\u65e5\u671f"))) for row in task_records})
    daily_rows = []
    for day_text in dates:
        day_tasks = [row for row in task_records if str(row.get(cn(r"\u65e5\u671f"))) == day_text]
        day_time = [row for row in time_records if str(row.get(cn(r"\u65e5\u671f"))) == day_text]
        day_evidence = [row for row in evidence_records if str(row.get(cn(r"\u65e5\u671f"))) == day_text]
        total_rows = [row for row in day_time if str(row.get(cn(r"\u5206\u7c7b")) or "") == cn(r"\u5408\u8ba1")]
        if total_rows:
            ai_minutes = parse_minutes(total_rows[0].get(cn(r"AI\u6d3b\u8dc3\u8017\u65f6")))
            work_minutes = parse_minutes(total_rows[0].get(cn(r"\u4f30\u7b97\u771f\u5b9e\u5de5\u4f5c\u8017\u65f6")))
        else:
            ai_minutes = sum(parse_minutes(row.get(cn(r"AI\u6d3b\u8dc3\u8017\u65f6"))) for row in day_tasks)
            work_minutes = sum(parse_minutes(row.get(cn(r"\u4f30\u7b97\u771f\u5b9e\u5de5\u4f5c\u8017\u65f6"))) for row in day_tasks)
        statuses = [str(row.get(cn(r"\u72b6\u6001")) or "") for row in day_tasks]
        strengths = [str(row.get(cn(r"\u8bc1\u636e\u5f3a\u5ea6")) or "") for row in day_tasks]
        projects = [str(row.get(cn(r"\u6d89\u53ca\u9879\u76ee")) or "") for row in day_tasks if row.get(cn(r"\u6d89\u53ca\u9879\u76ee"))]
        main_projects = cn(r"\u3001").join(dict.fromkeys(projects[:5]))
        commit_count = sum(1 for row in day_evidence if "commit" in str(row.get(cn(r"\u7c7b\u578b")) or "").lower())
        daily_rows.append({
            "date": day_text,
            "ai": ai_minutes,
            "work": work_minutes,
            "tasks": len(day_tasks),
            "commits": commit_count,
            "submitted": sum(1 for value in statuses if cn(r"\u5df2\u63d0\u4ea4") in value),
            "analyzed": sum(1 for value in statuses if cn(r"\u5df2\u5206\u6790") in value),
            "done": sum(1 for value in statuses if cn(r"\u5df2\u5b8c\u6210") in value),
            "strong": strengths.count(cn(r"\u5f3a")),
            "middle": strengths.count(cn(r"\u4e2d")),
            "weak": strengths.count(cn(r"\u5f31")),
            "projects": main_projects,
        })

    overview = reset_summary_sheet(cn(r"\u65e5\u62a5\u603b\u89c8"), [cn(r"\u65e5\u671f"), cn(r"AI\u6d3b\u8dc3\u8017\u65f6"), cn(r"\u4f30\u7b97\u771f\u5b9e\u5de5\u4f5c\u8017\u65f6"), cn(r"\u4efb\u52a1\u6570"), cn(r"Git\u63d0\u4ea4\u6570"), cn(r"\u5df2\u63d0\u4ea4"), cn(r"\u5df2\u5206\u6790"), cn(r"\u5df2\u5b8c\u6210"), cn(r"\u5f3a\u8bc1\u636e\u4efb\u52a1"), cn(r"\u4e2d\u8bc1\u636e\u4efb\u52a1"), cn(r"\u5f31\u8bc1\u636e\u4efb\u52a1"), cn(r"\u4e3b\u8981\u9879\u76ee")], 0)
    for row in daily_rows:
        overview.append([row["date"], format_duration(row["ai"]), format_duration(row["work"]), row["tasks"], row["commits"], row["submitted"], row["analyzed"], row["done"], row["strong"], row["middle"], row["weak"], row["projects"]])

    def period_rows(kind: str) -> list[list[Any]]:
        groups: dict[str, list[dict[str, Any]]] = {}
        for row in daily_rows:
            try:
                day = dt.date.fromisoformat(row["date"])
            except ValueError:
                continue
            if kind == "week":
                year, week, _ = day.isocalendar()
                key = f"{year}-W{week:02d}"
            else:
                key = day.strftime("%Y-%m")
            groups.setdefault(key, []).append(row)
        result = []
        for key, rows in sorted(groups.items()):
            result.append([
                key,
                f"{rows[0]['date']}~{rows[-1]['date']}",
                format_duration(sum(row["ai"] for row in rows)),
                format_duration(sum(row["work"] for row in rows)),
                sum(row["tasks"] for row in rows),
                sum(row["commits"] for row in rows),
                sum(row["submitted"] for row in rows),
                sum(row["analyzed"] for row in rows),
                sum(row["done"] for row in rows),
                sum(row["strong"] for row in rows),
                sum(row["middle"] for row in rows),
                sum(row["weak"] for row in rows),
            ])
        return result

    weekly = reset_summary_sheet(cn(r"\u5468\u6c47\u603b"), [cn(r"\u5468\u671f"), cn(r"\u65e5\u671f\u8303\u56f4"), cn(r"AI\u6d3b\u8dc3\u8017\u65f6"), cn(r"\u4f30\u7b97\u771f\u5b9e\u5de5\u4f5c\u8017\u65f6"), cn(r"\u4efb\u52a1\u6570"), cn(r"Git\u63d0\u4ea4\u6570"), cn(r"\u5df2\u63d0\u4ea4"), cn(r"\u5df2\u5206\u6790"), cn(r"\u5df2\u5b8c\u6210"), cn(r"\u5f3a\u8bc1\u636e\u4efb\u52a1"), cn(r"\u4e2d\u8bc1\u636e\u4efb\u52a1"), cn(r"\u5f31\u8bc1\u636e\u4efb\u52a1")], 1)
    for row in period_rows("week"):
        weekly.append(row)

    monthly = reset_summary_sheet(cn(r"\u6708\u6c47\u603b"), [cn(r"\u6708\u4efd"), cn(r"\u65e5\u671f\u8303\u56f4"), cn(r"AI\u6d3b\u8dc3\u8017\u65f6"), cn(r"\u4f30\u7b97\u771f\u5b9e\u5de5\u4f5c\u8017\u65f6"), cn(r"\u4efb\u52a1\u6570"), cn(r"Git\u63d0\u4ea4\u6570"), cn(r"\u5df2\u63d0\u4ea4"), cn(r"\u5df2\u5206\u6790"), cn(r"\u5df2\u5b8c\u6210"), cn(r"\u5f3a\u8bc1\u636e\u4efb\u52a1"), cn(r"\u4e2d\u8bc1\u636e\u4efb\u52a1"), cn(r"\u5f31\u8bc1\u636e\u4efb\u52a1")], 2)
    for row in period_rows("month"):
        monthly.append(row)

    for ws in wb.worksheets:
        if ws.max_row:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_len = max((len(str(cell.value or "")) for cell in list(column_cells)[:120]), default=0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 70)

    wb.save(output_path)
    return {"path": str(output_path), "mode": mode, "sheets": wb.sheetnames}


def main() -> None:
    parser = argparse.ArgumentParser(description="Collect Codex/Claude AI worklog evidence")
    parser.add_argument("--date", default="today", help="today, yesterday, or YYYY-MM-DD")
    parser.add_argument("--format", choices=["json", "markdown"], default="markdown")
    parser.add_argument("--root", action="append", default=[], help="Root directory to discover git repos")
    parser.add_argument("--repo", action="append", default=[], help="Specific git repo to inspect")
    parser.add_argument("--excel", nargs="?", const="", default=None, help="Append/upsert results to .xlsx. Default path: skill data/ai_worklog.xlsx")
    parser.add_argument("--excel-mode", choices=["upsert", "append"], default="upsert", help="Excel write mode. upsert removes existing rows for the same date first.")
    parser.add_argument("--preview", nargs="?", const="", default=None, help="Write privacy review preview .xlsx and stop. Default path: skill data/ai_worklog_preview.xlsx")
    parser.add_argument("--from-preview", default=None, help="Read confirmed rows from preview .xlsx before writing the final ledger")
    parser.add_argument("--config", default=None, help="Privacy config JSONC path. Default path: skill data/privacy_config.local.jsonc")
    parser.add_argument("--state", default=None, help="Dialogue state JSON path. Default path: skill data/ai_worklog_state.local.json")
    parser.add_argument("--from-state", action="store_true", help="Use the preview path recorded in the latest dialogue state")
    parser.add_argument("--include-keyword", action="append", default=[], help="Temporary whitelist keyword for this preview run")
    parser.add_argument("--exclude-keyword", action="append", default=[], help="Temporary privacy blacklist keyword for this preview run")
    parser.add_argument("--exclude-tool", action="append", default=[], help="Temporary AI tool exclusion, for example Claude")
    parser.add_argument("--exclude-time", action="append", default=[], help="Temporary excluded time range, format HH:mm-HH:mm")
    parser.add_argument("--policy", choices=["safe_only", "review_all", "include_all_except_high"], default=None, help="Temporary default include policy for this preview run")
    args = parser.parse_args()

    day = parse_date(args.date)
    codex_rows, codex_files, codex_cwds = collect_codex(day)
    claude_rows, claude_files, claude_cwds = collect_claude(day)
    repos = discover_repos(args.repo, args.root, codex_cwds | claude_cwds)
    git_rows = collect_git(day, repos)
    conversations = [row_to_jsonable(r) for r in sorted([*codex_rows, *claude_rows], key=lambda r: r.get("start") or dt.datetime.min.replace(tzinfo=dt.timezone.utc))]

    state_path = resolve_state_path(args.state)

    if args.preview is not None:
        config, config_path = load_privacy_config(args.config)
        config = apply_cli_filters(config, args)
        preview_path = resolve_preview_path(args.preview)
        preview_result = write_preview_excel(conversations, config, config_path, preview_path, day)
        summary = preview_summary(conversations, config)
        write_state(state_path, {
            "date": day.isoformat(),
            "preview_path": str(preview_path),
            "config_path": str(config_path),
            "filters": {
                "include_keyword": args.include_keyword,
                "exclude_keyword": args.exclude_keyword,
                "exclude_tool": args.exclude_tool,
                "exclude_time": args.exclude_time,
                "policy": args.policy,
            },
            "preview_summary": summary,
        })
        if args.format == "json":
            print(json.dumps({"date": day.isoformat(), "preview": preview_result, "summary": summary, "state": str(state_path)}, ensure_ascii=False, indent=2))
        else:
            print(cn(r"\u5df2\u751f\u6210\u5019\u9009\u7d20\u6750\u6e05\u5355\uff1a") + preview_result["path"])
            print(cn(r"\u672c\u5730\u9690\u79c1\u914d\u7f6e\uff1a") + preview_result["config"])
            print(cn(r"\u5bf9\u8bdd\u72b6\u6001\uff1a") + str(state_path))
        return

    if args.from_state and not args.from_preview:
        state = read_state(state_path)
        args.from_preview = state.get("preview_path")
        if not args.from_preview:
            raise SystemExit(cn(r"\u672a\u627e\u5230\u4e0a\u4e00\u6b21 preview 状\u6001\uff0c\u8bf7\u5148\u751f\u6210 preview"))

    if args.from_preview:
        conversations = read_confirmed_preview(Path(args.from_preview).expanduser().resolve())

    data = {
        "date": day.isoformat(),
        "sources": [
            {"source": "Codex", "path": str(codex_home()), "count": len(codex_files), "status": "found" if codex_files else "none"},
            {"source": "Claude", "path": str(claude_home()), "count": len(claude_files), "status": "found" if claude_files else "none"},
            {"source": "Git", "path": ", ".join(str(r) for r in repos), "count": len(repos), "status": "found" if repos else "none"},
        ],
        "conversations": conversations,
        "git": git_rows,
    }
    excel_result = None
    if args.excel is not None:
        excel_result = write_excel(data, resolve_excel_path(args.excel), args.excel_mode)
        data["excel"] = excel_result

    if args.format == "json":
        print(json.dumps(data, ensure_ascii=False, indent=2))
    else:
        print(render_markdown(data))
        if excel_result:
            print(f"\n## Excel 台账\n已写入：{excel_result['path']}")


if __name__ == "__main__":
    main()
